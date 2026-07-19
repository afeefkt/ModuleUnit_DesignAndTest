"""Excel (.xlsx) requirement importer using openpyxl.

Supports Polarion and DOORS export formats with configurable column mapping.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

from mudtool.importers.base import BaseImporter, ImportResult
from mudtool.models.requirements import RequirementSet

logger = logging.getLogger(__name__)


class ExcelImporter(BaseImporter):
    """Import requirements from Excel (.xlsx) files.

    Supports:
    - Polarion ALM Excel exports
    - IBM DOORS Excel exports
    - Custom Excel templates with configurable column mapping
    """

    def __init__(
        self,
        column_mapping: Optional[dict[str, str]] = None,
        sheet_name: Optional[str] = None,
        header_row: int = 1,
    ):
        super().__init__(column_mapping)
        self.sheet_name = sheet_name  # None = active sheet
        self.header_row = header_row

    def supports_format(self, file_path: Path) -> bool:
        return file_path.suffix.lower() in (".xlsx", ".xls")

    def import_file(self, file_path: Path) -> ImportResult:
        """Import requirements from an Excel file.

        The importer:
        1. Opens the workbook (specified sheet or active sheet)
        2. Reads headers from header_row
        3. Auto-detects column mapping from headers
        4. Iterates data rows and builds Requirement objects
        """
        import openpyxl

        self._warnings = []
        self._errors = []
        requirements = []
        rows_processed = 0
        rows_skipped = 0

        try:
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        except Exception as e:
            return ImportResult(
                requirement_set=RequirementSet(source_file=str(file_path), source_format="xlsx"),
                errors=[f"Failed to open Excel file: {e}"],
            )

        try:
            # Select sheet
            if self.sheet_name:
                if self.sheet_name not in wb.sheetnames:
                    return ImportResult(
                        requirement_set=RequirementSet(
                            source_file=str(file_path), source_format="xlsx"
                        ),
                        errors=[
                            f"Sheet '{self.sheet_name}' not found. "
                            f"Available: {wb.sheetnames}"
                        ],
                    )
                ws = wb[self.sheet_name]
            else:
                ws = wb.active

            if ws is None:
                return ImportResult(
                    requirement_set=RequirementSet(
                        source_file=str(file_path), source_format="xlsx"
                    ),
                    errors=["No active worksheet found"],
                )

            # Auto-detect header row: scan first 30 rows, pick the row that has
            # req_id (or the most alias matches overall, minimum 2).
            from mudtool.importers.base import DEFAULT_COLUMN_MAPPING
            all_aliases = {
                alias.lower(): field
                for field, aliases in DEFAULT_COLUMN_MAPPING.items()
                for alias in aliases
            }
            req_id_aliases = {a.lower() for a in DEFAULT_COLUMN_MAPPING.get("req_id", [])}

            best_row = self.header_row
            best_headers: list[str] = []
            best_score = -1

            for scan_row in range(1, 31):
                try:
                    candidate = [str(cell.value or "").strip() for cell in ws[scan_row]]
                except Exception:
                    break
                has_req_id = any(h.lower() in req_id_aliases for h in candidate if h)
                match_count = sum(1 for h in candidate if h and h.lower() in all_aliases)
                # Prefer rows that have req_id; break immediately when found
                if has_req_id:
                    best_row = scan_row
                    best_headers = candidate
                    break
                # Otherwise track the row with most matches (min 2)
                if match_count >= 2 and match_count > best_score:
                    best_score = match_count
                    best_row = scan_row
                    best_headers = candidate

            actual_header_row = best_row
            headers = best_headers if best_headers else [
                str(cell.value or "").strip() for cell in ws[self.header_row]
            ]
            if actual_header_row != self.header_row:
                logger.info(
                    "Auto-detected header row %d (scanned up to row 30)",
                    actual_header_row,
                )

            # Build column index mapping: col_index -> canonical field name
            col_map: dict[int, str] = {}
            unmapped_cols: list[str] = []

            for idx, header in enumerate(headers):
                if not header:
                    continue
                field = self._detect_column(header)
                if field:
                    col_map[idx] = field
                else:
                    unmapped_cols.append(header)

            if unmapped_cols:
                self._warnings.append(f"Unmapped columns (ignored): {unmapped_cols}")

            # Validate required columns
            mapped_fields = set(col_map.values())
            if "req_id" not in mapped_fields:
                return ImportResult(
                    requirement_set=RequirementSet(
                        source_file=str(file_path), source_format="xlsx"
                    ),
                    errors=[
                        "Required column 'Req_ID' not found in headers. "
                        f"Found: {headers}"
                    ],
                )

            # Read data rows (start after the auto-detected header row)
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=actual_header_row + 1), start=actual_header_row + 1
            ):
                rows_processed += 1
                row_data: dict[str, str] = {}

                for col_idx, cell in enumerate(row):
                    if col_idx in col_map:
                        value = cell.value
                        row_data[col_map[col_idx]] = str(value).strip() if value else ""

                # Skip entirely empty rows
                if not any(v.strip() for v in row_data.values()):
                    rows_skipped += 1
                    continue

                req = self._build_requirement(row_data, row_idx)
                if req:
                    requirements.append(req)
                else:
                    rows_skipped += 1

        finally:
            wb.close()

        req_set = RequirementSet(
            requirements=requirements,
            source_file=str(file_path),
            source_format="xlsx",
            column_mapping={v: headers[k] for k, v in col_map.items() if k < len(headers)},
        )

        logger.info(
            f"Imported {len(requirements)} requirements from {file_path.name} "
            f"({rows_processed} rows, {rows_skipped} skipped)"
        )

        return ImportResult(
            requirement_set=req_set,
            warnings=self._warnings,
            errors=self._errors,
            rows_processed=rows_processed,
            rows_skipped=rows_skipped,
        )
