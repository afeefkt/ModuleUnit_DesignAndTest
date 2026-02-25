"""Excel (.xlsx) requirement importer using openpyxl.

Supports Polarion and DOORS export formats with configurable column mapping.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

from mudtool.importers.base import BaseImporter, ImportResult
from mudtool.models.requirements import RequirementSet

logger = logging.getLogger(__name__)


class ExcelImporter(BaseImporter):
    """Import requirements from Excel (.xlsx) files.

    Supports:
    - Polarion ALM Excel exports
    - IBM DOORS Excel exports
    - Custom Excel templates with configurable column mapping
    """

    def __init__(
        self,
        column_mapping: Optional[dict[str, str]] = None,
        sheet_name: Optional[str] = None,
        header_row: int = 1,
    ):
        super().__init__(column_mapping)
        self.sheet_name = sheet_name  # None = active sheet
        self.header_row = header_row

    def supports_format(self, file_path: Path) -> bool:
        return file_path.suffix.lower() in (".xlsx", ".xls")

    def import_file(self, file_path: Path) -> ImportResult:
        """Import requirements from an Excel file.

        The importer:
        1. Opens the workbook (specified sheet or active sheet)
        2. Reads headers from header_row
        3. Auto-detects column mapping from headers
        4. Iterates data rows and builds Requirement objects
        """
        import openpyxl

        self._warnings = []
        self._errors = []
        requirements = []
        rows_processed = 0
        rows_skipped = 0

        try:
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        except Exception as e:
            return ImportResult(
                requirement_set=RequirementSet(source_file=str(file_path), source_format="xlsx"),
                errors=[f"Failed to open Excel file: {e}"],
            )

        try:
            # Select sheet
            if self.sheet_name:
                if self.sheet_name not in wb.sheetnames:
                    return ImportResult(
                        requirement_set=RequirementSet(
                            source_file=str(file_path), source_format="xlsx"
                        ),
                        errors=[
                            f"Sheet '{self.sheet_name}' not found. "
                            f"Available: {wb.sheetnames}"
                        ],
                    )
                ws = wb[self.sheet_name]
            else:
                ws = wb.active

            if ws is None:
                return ImportResult(
                    requirement_set=RequirementSet(
                        source_file=str(file_path), source_format="xlsx"
                    ),
                    errors=["No active worksheet found"],
                )

            # Read headers
            headers: list[str] = []
            for cell in ws[self.header_row]:
                headers.append(str(cell.value or "").strip())

            # Build column index mapping: col_index -> canonical field name
            col_map: dict[int, str] = {}
            unmapped_cols: list[str] = []

            for idx, header in enumerate(headers):
                if not header:
                    continue
                field = self._detect_column(header)
                if field:
                    col_map[idx] = field
                else:
                    unmapped_cols.append(header)

            if unmapped_cols:
                self._warnings.append(f"Unmapped columns (ignored): {unmapped_cols}")

            # Validate required columns
            mapped_fields = set(col_map.values())
            if "req_id" not in mapped_fields:
                return ImportResult(
                    requirement_set=RequirementSet(
                        source_file=str(file_path), source_format="xlsx"
                    ),
                    errors=[
                        "Required column 'Req_ID' not found in headers. "
                        f"Found: {headers}"
                    ],
                )

            # Read data rows
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=self.header_row + 1), start=self.header_row + 1
            ):
                rows_processed += 1
                row_data: dict[str, str] = {}

                for col_idx, cell in enumerate(row):
                    if col_idx in col_map:
                        value = cell.value
                        row_data[col_map[col_idx]] = str(value).strip() if value else ""

                # Skip entirely empty rows
                if not any(v.strip() for v in row_data.values()):
                    rows_skipped += 1
                    continue

                req = self._build_requirement(row_data, row_idx)
                if req:
                    requirements.append(req)
                else:
                    rows_skipped += 1

        finally:
            wb.close()

        req_set = RequirementSet(
            requirements=requirements,
            source_file=str(file_path),
            source_format="xlsx",
            column_mapping={v: headers[k] for k, v in col_map.items() if k < len(headers)},
        )

        logger.info(
            f"Imported {len(requirements)} requirements from {file_path.name} "
            f"({rows_processed} rows, {rows_skipped} skipped)"
        )

        return ImportResult(
            requirement_set=req_set,
            warnings=self._warnings,
            errors=self._errors,
            rows_processed=rows_processed,
            rows_skipped=rows_skipped,
        )
